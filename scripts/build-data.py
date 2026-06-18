from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "商品.xlsx"
QR_DIR = ROOT / "mini_qrcode_export"
OUTPUT = ROOT / "public" / "data" / "products.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value):
    if value is None or value == "":
        return None
    return float(value)


def display_number(value):
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value


def load_qr_files():
    qr_by_product = {}
    for file in QR_DIR.iterdir():
        if not file.is_file():
            continue
        match = re.search(r"_(\d+)_", file.name)
        if not match:
            continue
        product_id = match.group(1)
        qr_by_product[product_id] = {
            "fileName": file.name,
            "path": "/mini_qrcode_export/" + quote(file.name),
        }
    return qr_by_product


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    qr_by_product = load_qr_files()
    products = OrderedDict()

    for values in rows:
        if not any(value is not None for value in values):
            continue
        row = dict(zip(headers, values))
        product_id = clean(row.get("商品Id"))
        if not product_id:
            continue

        price = number(row.get("销售价"))
        spec_value = clean(row.get("商品规格值"))
        sku_id = clean(row.get("商品SKUId"))
        sku_code = clean(row.get("SKU编码"))

        if product_id not in products:
            qr = qr_by_product.get(product_id)
            products[product_id] = {
                "productId": product_id,
                "name": clean(row.get("商品名称")),
                "merchant": clean(row.get("商家")),
                "brand": clean(row.get("品牌名称")) or "其他",
                "category1": clean(row.get("一级分类")),
                "category2": clean(row.get("二级分类")),
                "category3": clean(row.get("三级分类")),
                "priceMin": price,
                "priceMax": price,
                "specs": [],
                "qrCodePath": qr["path"] if qr else "",
                "qrFileName": qr["fileName"] if qr else "",
                "hasQrCode": bool(qr),
            }

        product = products[product_id]
        if price is not None:
            product["priceMin"] = price if product["priceMin"] is None else min(product["priceMin"], price)
            product["priceMax"] = price if product["priceMax"] is None else max(product["priceMax"], price)
        product["specs"].append(
            {
                "skuId": sku_id,
                "skuCode": sku_code,
                "value": spec_value,
                "price": price,
            }
        )

    output = []
    for product in products.values():
        searchable_parts = [
            product["name"],
            product["merchant"],
            product["brand"],
            product["category1"],
            product["category2"],
            product["category3"],
            " ".join(spec["value"] for spec in product["specs"]),
        ]
        product["priceMin"] = display_number(product["priceMin"])
        product["priceMax"] = display_number(product["priceMax"])
        for spec in product["specs"]:
            spec["price"] = display_number(spec["price"])
        product["searchText"] = " ".join(part for part in searchable_parts if part)
        output.append(product)

    output.sort(key=lambda item: (not item["hasQrCode"], item["category2"], item["brand"], item["name"]))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(output)} products to {OUTPUT}")
    print(f"Products with QR code: {sum(1 for item in output if item['hasQrCode'])}")


if __name__ == "__main__":
    main()
